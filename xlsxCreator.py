#!/usr/bin/python
 # -*- coding: utf-8 -*-
import os, platform
from string import uppercase as UC
from email.mime.base import MIMEBase
from openpyxl import Workbook as W
from email import encoders 
div = '\\' if platform.system() is 'Windows' else '/'
class Excel():
    def __init__(self,fileName,workSheetName,matrix,headers):
        self.fileName = fileName
        self.matrix = matrix
        self.headers = headers
        self.workbook = W() #Create file
        self.firstSheet=True
        self.generateWorksheet(workSheetName,matrix,headers)
    def __enter__(self):
        return self
    def __exit__(self, exc_type, exc_value, traceback):
        self.DeleteFile()
    def __del__(self):
        pass
    def generateWorksheet(self,workSheetName,matrix,headers):
        ws =  self.workbook.active if self.firstSheet else self.workbook.create_sheet()
        self.firstSheet=False
        ws.title = workSheetName
        x=y=1 
        for header in headers:      #WritHeadersROW1
            ws['%s1'%UC[x]]=header
            x+=1
        x=1
        for row in matrix:
            y+=1
            ws['A%s'%y]=y-1           #RowNumberAlwaysColumnA
            for cell in row:
                ws['%s%s'%(UC[x],y)]=cell #WriteCell
                x+=1
            x=1
        self.workbook.save(self.fileName)
    def getFilePath(self):
        return os.getcwd()+div+self.fileName
    def DeleteFile(self):
        os.remove(self.fileName)
    def getAsMailAttachment(self):
        part = MIMEBase('application', "octet-stream")
        part.set_payload(open(self.fileName, "rb").read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment; filename="%s"'%self.fileName)
        return part